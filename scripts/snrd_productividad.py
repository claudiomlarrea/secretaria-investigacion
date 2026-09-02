#!/usr/bin/env python3
"""
Cruce preliminar SNRD — UCCuyo.

Combina OpenAlex (afiliación institucional), planilla de publicaciones del sitio
(Apps Script) y, opcionalmente, un export del RI con acceso abierto confirmado.

Genera:
  - Excel SNRD prellenado (filas automatizables)
  - CSV de detalle para revisión fila por fila

Filas dejadas en 0 (completar con Rectorado / otras fuentes):
  Tesis, Revistas editadas por la institución, Patentes, Informes técnicos.

Uso:
  python3 scripts/snrd_productividad.py \\
    --template ~/Downloads/Productividad_UCCuyo_SNRD.xlsx \\
    --out ~/Downloads/Productividad_UCCuyo_SNRD_v1.xlsx

  python3 scripts/snrd_productividad.py --ri ~/Downloads/ri_acceso_abierto.csv
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalá openpyxl: pip install openpyxl", file=sys.stderr)
    raise

INSTITUTION_ID = "I4210121591"
MAILTO = "investigacion@uccuyo.edu.ar"
APPS_SCRIPT_URL = (
    "https://script.google.com/a/macros/uccuyo.edu.ar/s/"
    "AKfycbxXOx3XpKzmpffpUFJ9tLctA5FR-552RbggS4pLO2KrL3mpVVZuKyGBFdnXDC3qR5zH/exec"
)
DEFAULT_YEARS = (2022, 2023, 2024, 2025)
PER_PAGE = 100
SLEEP_SEC = 0.35

SNRD_ROWS = {
    "articulos": 10,
    "tesis": 11,
    "conferencias": 12,
    "libros": 13,
    "partes_libros": 14,
    "revistas_institucion": 15,
    "patentes": 16,
    "informes": 17,
}

YEAR_COLS = {
    2025: (2, 3, 4),
    2024: (5, 6, 7),
    2023: (8, 9, 10),
    2022: (11, 12, 13),
}

DETAIL_HEADERS = [
    "anio",
    "categoria_snrd",
    "titulo",
    "autores",
    "doi",
    "fuente",
    "tipo_origen",
    "digital",
    "acceso_abierto",
    "notas",
]


@dataclass
class Registro:
    anio: int
    categoria: str
    titulo: str
    autores: str = ""
    doi: str = ""
    fuente: str = ""
    tipo_origen: str = ""
    digital: bool = False
    oa: bool = False
    notas: str = ""
    clave: str = ""

    def __post_init__(self) -> None:
        if not self.clave:
            self.clave = dedupe_key(self.doi, self.titulo, self.anio)


def dedupe_key(doi: str, titulo: str, anio: int | str) -> str:
    d = normalize_doi(doi)
    if d:
        return f"doi:{d.lower()}"
    t = " ".join(str(titulo or "").lower().split())
    return f"t:{t}|y:{anio}"


def normalize_doi(raw: str) -> str:
    s = (raw or "").strip()
    for prefix in ("https://doi.org/", "http://doi.org/", "https://dx.doi.org/", "doi:"):
        if s.lower().startswith(prefix):
            s = s[len(prefix) :].strip()
    return s


def fetch_json(url: str) -> dict:
    req = urllib.request.Request(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": f"mailto:{MAILTO} (snrd-productividad)",
        },
    )
    for attempt in range(1, 8):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < 7:
                time.sleep(2**attempt)
                continue
            raise
    raise RuntimeError("Sin respuesta tras reintentos")


def fetch_openalex_year(year: int) -> list[dict]:
    page = 1
    out: list[dict] = []
    filters = f"authorships.institutions.lineage:{INSTITUTION_ID},publication_year:{year}"
    while True:
        params = {
            "filter": filters,
            "sort": "publication_date:desc",
            "per-page": str(PER_PAGE),
            "page": str(page),
            "mailto": MAILTO,
        }
        url = "https://api.openalex.org/works?" + urllib.parse.urlencode(params)
        print(f"  OpenAlex {year}, pág. {page}…", file=sys.stderr)
        data = fetch_json(url)
        results = data.get("results") or []
        if not results:
            break
        out.extend(results)
        if len(results) < PER_PAGE:
            break
        page += 1
        time.sleep(SLEEP_SEC)
    return out


def openalex_categoria(work: dict) -> str | None:
    t = (work.get("type") or "").lower()
    if t in {"article", "review"}:
        return "articulos"
    if t in {"conference-paper", "proceedings"}:
        return "conferencias"
    if t == "book":
        return "libros"
    if t == "book-chapter":
        return "partes_libros"
    if t in {"dissertation"}:
        return "tesis"
    if t in {"report"}:
        return "informes"
    return None


def work_es_digital(work: dict) -> bool:
    if normalize_doi(work.get("doi") or ""):
        return True
    loc = work.get("primary_location") or {}
    if loc.get("landing_page_url") or loc.get("pdf_url"):
        return True
    for loc_any in work.get("locations") or []:
        if loc_any.get("landing_page_url") or loc_any.get("pdf_url"):
            return True
    return False


def work_es_oa(work: dict) -> bool:
    oa = work.get("open_access") or {}
    if oa.get("is_oa"):
        return True
    status = (oa.get("oa_status") or "").lower()
    return status in {"gold", "green", "hybrid", "bronze", "diamond"}


def openalex_a_registro(work: dict) -> Registro | None:
    year = work.get("publication_year")
    if not year or year not in DEFAULT_YEARS:
        return None
    cat = openalex_categoria(work)
    if not cat:
        return None
    if cat == "tesis":
        return None  # Rectorado
    if cat == "informes":
        return None  # Rectorado

    titulo = (work.get("display_name") or "").strip()
    if not titulo:
        return None

    autores = ", ".join(
        a.get("author", {}).get("display_name", "")
        for a in work.get("authorships") or []
        if a.get("author", {}).get("display_name")
    )
    doi = normalize_doi(work.get("doi") or "")
    notas = f"OpenAlex type={work.get('type')}"
    if (work.get("type") or "").lower() == "review":
        notas += " · revisar si cuenta como artículo arbitrado"

    return Registro(
        anio=int(year),
        categoria=cat,
        titulo=titulo,
        autores=autores,
        doi=doi,
        fuente="OpenAlex",
        tipo_origen=str(work.get("type") or ""),
        digital=work_es_digital(work),
        oa=work_es_oa(work),
        notas=notas,
    )


def fetch_planilla_publicaciones() -> list[dict]:
    print("  Planilla publicaciones (Apps Script)…", file=sys.stderr)
    data = fetch_json(APPS_SCRIPT_URL)
    if not data.get("ok"):
        raise RuntimeError("Respuesta inválida del Apps Script de publicaciones")
    return data.get("items") or []


def planilla_categoria(item: dict) -> str | None:
    cat = (item.get("categoria") or "").lower().strip()
    tipo = (item.get("tipo_origen") or "").lower().strip()
    tp = (item.get("tipo_publicacion") or "").lower().strip()

    if cat == "revistas" or tipo == "revista":
        return "articulos"
    if cat == "eventos" or tipo == "evento":
        return "conferencias"
    if cat == "libros" or tipo == "libro":
        if "capitulo" in tp or "capítulo" in tp:
            return "partes_libros"
        return "libros"
    if "capitulo" in tipo or "capítulo" in tipo or "capitulo" in tp or "capítulo" in tp:
        return "partes_libros"
    if cat == "repositorios" or tipo == "repositorio":
        return None  # Rectorado / informes
    if cat == "diarios" or tipo == "diario":
        return None  # fuera del formulario SNRD
    return None


def planilla_a_registro(item: dict) -> Registro | None:
    anio_raw = str(item.get("anio") or item.get("fecha") or "")[:4]
    if not anio_raw.isdigit():
        return None
    anio = int(anio_raw)
    if anio not in DEFAULT_YEARS:
        return None

    cat = planilla_categoria(item)
    if not cat:
        return None

    titulo = (item.get("titulo") or "").strip()
    if not titulo:
        return None

    doi = normalize_doi(item.get("doi") or "")
    link = (item.get("link") or "").strip()
    digital = bool(doi or link)

    return Registro(
        anio=anio,
        categoria=cat,
        titulo=titulo,
        autores=(item.get("autores") or "").strip(),
        doi=doi,
        fuente="Planilla web",
        tipo_origen=(item.get("tipo_origen") or item.get("categoria") or "").strip(),
        digital=digital,
        oa=False,
        notas="OA no confirmado en planilla; revisar manualmente",
    )


def cargar_ri(path: Path) -> list[dict]:
    rows: list[dict] = []
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))
    elif suffix in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        wb.close()
    else:
        raise ValueError(f"Formato RI no soportado: {suffix}")
    print(f"  RI: {len(rows)} filas desde {path.name}", file=sys.stderr)
    return rows


def ri_enriquecer(registros: dict[str, Registro], ri_rows: list[dict]) -> None:
    """Marca OA confirmado en RI por DOI; agrega registros RI-only si aplica."""
    doi_oa: set[str] = set()
    for row in ri_rows:
        doi = normalize_doi(str(row.get("doi") or row.get("DOI") or ""))
        if not doi:
            continue
        oa_flag = str(row.get("acceso_abierto") or row.get("oa") or row.get("open_access") or "").lower()
        if oa_flag in {"1", "true", "si", "sí", "yes", "x"}:
            doi_oa.add(doi.lower())

    for reg in registros.values():
        if reg.doi and reg.doi.lower() in doi_oa:
            reg.oa = True
            reg.notas = (reg.notas + " · OA confirmado RI").strip(" ·")


def fusionar(registros: list[Registro]) -> dict[str, Registro]:
    merged: dict[str, Registro] = {}
    for reg in registros:
        prev = merged.get(reg.clave)
        if not prev:
            merged[reg.clave] = reg
            continue
        if not prev.doi and reg.doi:
            prev.doi = reg.doi
        if not prev.digital and reg.digital:
            prev.digital = True
        if not prev.oa and reg.oa:
            prev.oa = True
        if reg.fuente not in prev.fuente:
            prev.fuente = prev.fuente + " + " + reg.fuente if prev.fuente else reg.fuente
        prev.notas = " · ".join(filter(None, {prev.notas, reg.notas}))
    return merged


def contar(registros: dict[str, Registro]) -> dict[str, dict[int, dict[str, int]]]:
    counts: dict[str, dict[int, dict[str, int]]] = defaultdict(
        lambda: defaultdict(lambda: {"total": 0, "digital": 0, "oa": 0})
    )
    for reg in registros.values():
        bucket = counts[reg.categoria][reg.anio]
        bucket["total"] += 1
        if reg.digital:
            bucket["digital"] += 1
        if reg.oa:
            bucket["oa"] += 1
    return counts


def escribir_excel(template: Path, out: Path, counts: dict) -> None:
    wb = openpyxl.load_workbook(template)
    ws = wb["Productividad"]
    for cat, row in SNRD_ROWS.items():
        if cat in {"tesis", "revistas_institucion", "patentes", "informes"}:
            for year, (c_total, c_digital, c_oa) in YEAR_COLS.items():
                ws.cell(row=row, column=c_total, value=0)
                ws.cell(row=row, column=c_digital, value=0)
                ws.cell(row=row, column=c_oa, value=0)
            continue
        for year, (c_total, c_digital, c_oa) in YEAR_COLS.items():
            bucket = counts.get(cat, {}).get(year, {"total": 0, "digital": 0, "oa": 0})
            ws.cell(row=row, column=c_total, value=bucket["total"])
            ws.cell(row=row, column=c_digital, value=bucket["digital"])
            ws.cell(row=row, column=c_oa, value=bucket["oa"])
    wb.save(out)
    wb.close()


def escribir_detalle(out: Path, registros: dict[str, Registro]) -> None:
    rows = sorted(
        registros.values(),
        key=lambda r: (-r.anio, r.categoria, r.titulo.lower()),
    )
    with out.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=DETAIL_HEADERS)
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "anio": r.anio,
                    "categoria_snrd": r.categoria,
                    "titulo": r.titulo,
                    "autores": r.autores,
                    "doi": r.doi,
                    "fuente": r.fuente,
                    "tipo_origen": r.tipo_origen,
                    "digital": int(r.digital),
                    "acceso_abierto": int(r.oa),
                    "notas": r.notas,
                }
            )


def imprimir_resumen(counts: dict, registros: dict[str, Registro]) -> None:
    print("\n=== Resumen preliminar (revisar antes de declarar) ===", file=sys.stderr)
    labels = {
        "articulos": "Artículos arbitrados",
        "conferencias": "Conferencias",
        "libros": "Libros",
        "partes_libros": "Partes de libros",
    }
    for cat, label in labels.items():
        print(f"\n{label}:", file=sys.stderr)
        for year in DEFAULT_YEARS:
            b = counts.get(cat, {}).get(year, {"total": 0, "digital": 0, "oa": 0})
            print(
                f"  {year}: total={b['total']} · digital={b['digital']} · OA={b['oa']}",
                file=sys.stderr,
            )
    print(
        f"\nRegistros únicos en detalle: {len(registros)}",
        file=sys.stderr,
    )
    print(
        "Filas en 0 (Rectorado): Tesis, Revistas institucionales, Patentes, Informes técnicos",
        file=sys.stderr,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Cruce preliminar SNRD UCCuyo")
    parser.add_argument(
        "--template",
        type=Path,
        default=Path.home() / "Downloads" / "Productividad_UCCuyo_SNRD.xlsx",
        help="Planilla SNRD vacía (plantilla)",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path.home() / "Downloads" / "Productividad_UCCuyo_SNRD_v1.xlsx",
        help="Excel SNRD prellenado",
    )
    parser.add_argument(
        "--detalle",
        type=Path,
        default=None,
        help="CSV de detalle (default: mismo nombre que --out con _detalle.csv)",
    )
    parser.add_argument(
        "--ri",
        type=Path,
        default=None,
        help="Export RI opcional (CSV/XLSX) con columna doi y acceso_abierto",
    )
    parser.add_argument(
        "--years",
        default="2022-2025",
        help="Rango de años, ej. 2022-2025",
    )
    args = parser.parse_args()

    if not args.template.is_file():
        print(f"No encontré plantilla: {args.template}", file=sys.stderr)
        return 1

    detalle_path = args.detalle or args.out.with_name(args.out.stem + "_detalle.csv")

    todos: list[Registro] = []

    print("Descargando OpenAlex…", file=sys.stderr)
    for year in DEFAULT_YEARS:
        for work in fetch_openalex_year(year):
            reg = openalex_a_registro(work)
            if reg:
                todos.append(reg)

    print("Descargando planilla de publicaciones…", file=sys.stderr)
    for item in fetch_planilla_publicaciones():
        reg = planilla_a_registro(item)
        if reg:
            todos.append(reg)

    merged = fusionar(todos)

    if args.ri and args.ri.is_file():
        ri_enriquecer(merged, cargar_ri(args.ri))

    counts = contar(merged)
    escribir_excel(args.template, args.out, counts)
    escribir_detalle(detalle_path, merged)
    imprimir_resumen(counts, merged)

    print(f"\nExcel:   {args.out}", file=sys.stderr)
    print(f"Detalle: {detalle_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
